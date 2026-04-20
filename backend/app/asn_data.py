import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PRIVATE_ASN_START = 64512
PRIVATE_ASN_END = 65534
TEXT_FILE_SUFFIX = ".txt"
EXCEL_FILE_SUFFIX = ".xlsx"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_asn(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None

    match = re.search(r"\d+", text.replace(",", ""))
    if not match:
        return None

    return int(match.group())


def parse_range(value: Any) -> tuple[int, int] | None:
    text = clean_text(value)
    match = re.search(r"(\d+)\s*-\s*(\d+)", text)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


class AsnRepository:
    def __init__(self, source_dir: str | None = None) -> None:
        self.source_dir = Path(source_dir or os.getenv("ASN_SOURCE_DIR", "/asn-source"))
        self._cache_signature: tuple[tuple[str, int, int], ...] | None = None
        self._cache_data: dict[str, Any] | None = None

    def load(self) -> dict[str, Any]:
        source_files = self._collect_source_files()
        signature = tuple(
            sorted(
                (
                    str(path),
                    int(path.stat().st_mtime),
                    path.stat().st_size,
                )
                for path in source_files
            )
        )
        if self._cache_data is not None and signature == self._cache_signature:
            return self._cache_data

        data = self._build_dataset(source_files)
        self._cache_signature = signature
        self._cache_data = data
        return data

    def _collect_source_files(self) -> list[Path]:
        if not self.source_dir.exists():
            return []
        return [
            path
            for path in self.source_dir.iterdir()
            if path.is_file() and not path.name.startswith(".")
        ]

    def _build_dataset(self, source_files: list[Path]) -> dict[str, Any]:
        assignments: dict[int, dict[str, Any]] = {}
        allocation_blocks: dict[tuple[int, int, str], dict[str, Any]] = {}
        route_snapshots: list[dict[str, Any]] = []
        warnings: list[str] = []

        for path in source_files:
            try:
                if path.suffix.lower() == EXCEL_FILE_SUFFIX:
                    self._load_excel(path, assignments, allocation_blocks)
                elif path.suffix.lower() == TEXT_FILE_SUFFIX:
                    route_snapshots.append(self._load_route_snapshot(path))
            except Exception as exc:
                warnings.append(f"{path.name}: {exc}")

        used_private = sorted(
            asn for asn in assignments if PRIVATE_ASN_START <= asn <= PRIVATE_ASN_END
        )
        available_private = [
            asn
            for asn in range(PRIVATE_ASN_START, PRIVATE_ASN_END + 1)
            if asn not in assignments
        ]
        duplicate_asns = sorted(
            (
                self._serialize_assignment(record)
                for record in assignments.values()
                if len(record["sources"]) > 1 or len(record["descriptions"]) > 1
            ),
            key=lambda item: item["asn"],
        )

        blocks = sorted(allocation_blocks.values(), key=lambda item: (item["start"], item["stop"]))
        recommendations = self._build_recommendations(blocks, assignments)
        serialized_assignments = sorted(
            (self._serialize_assignment(record) for record in assignments.values()),
            key=lambda item: item["asn"],
        )

        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_dir": str(self.source_dir),
            "source_files": sorted(path.name for path in source_files),
            "summary": {
                "used_private_count": len(used_private),
                "available_private_count": len(available_private),
                "duplicate_count": len(duplicate_asns),
                "route_snapshot_count": len(route_snapshots),
                "total_private_capacity": PRIVATE_ASN_END - PRIVATE_ASN_START + 1,
                "utilization_percent": round(
                    len(used_private) / (PRIVATE_ASN_END - PRIVATE_ASN_START + 1) * 100,
                    2,
                ),
                "next_available_asns": available_private[:10],
            },
            "blocks": blocks,
            "recommendations": recommendations,
            "duplicate_asns": duplicate_asns[:25],
            "route_snapshots": route_snapshots,
            "assignments": serialized_assignments,
            "warnings": warnings,
        }

    def _load_excel(
        self,
        path: Path,
        assignments: dict[int, dict[str, Any]],
        allocation_blocks: dict[tuple[int, int, str], dict[str, Any]],
    ) -> None:
        workbook = load_workbook(path, read_only=True, data_only=True)

        if path.name == "ASN-Assigment.master.xlsx":
            self._parse_master_workbook(workbook, assignments, allocation_blocks, path.name)
            return

        if path.name == "BGP AS Number at Ex.True-IT.xlsx":
            self._parse_external_workbook(workbook, assignments, path.name)
            return

        if path.name == "DTAC ACI Master-sheet-v5_update_20250901.xlsx":
            self._parse_aci_workbook(workbook, assignments, path.name)

    def _parse_master_workbook(
        self,
        workbook: Any,
        assignments: dict[int, dict[str, Any]],
        allocation_blocks: dict[tuple[int, int, str], dict[str, Any]],
        source: str,
    ) -> None:
        if "rawdata_dtac2" in workbook.sheetnames:
            sheet = workbook["rawdata_dtac2"]
            for row in sheet.iter_rows(min_row=4, values_only=True):
                for range_index, desc_index in ((0, 1), (4, 5)):
                    block_range = parse_range(row[range_index] if len(row) > range_index else None)
                    label = clean_text(row[desc_index] if len(row) > desc_index else None)
                    if block_range is None:
                        continue
                    start, stop = block_range
                    key = (start, stop, label or f"{start}-{stop}")
                    allocation_blocks[key] = {
                        "name": label or f"{start}-{stop}",
                        "start": start,
                        "stop": stop,
                        "source": source,
                    }

        if "rawdata_TrueIT" in workbook.sheetnames:
            sheet = workbook["rawdata_TrueIT"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self._record_assignment(
                    assignments,
                    parse_asn(row[0] if len(row) > 0 else None),
                    source,
                    description=clean_text(row[2] if len(row) > 2 else None),
                    tenant=clean_text(row[3] if len(row) > 3 else None),
                    domain="true-it",
                )
                self._record_assignment(
                    assignments,
                    parse_asn(row[6] if len(row) > 6 else None),
                    source,
                    description=clean_text(row[8] if len(row) > 8 else None),
                    tenant=clean_text(row[9] if len(row) > 9 else None),
                    domain=clean_text(row[7] if len(row) > 7 else None) or "true-it",
                )

        if "rawdata_true_tx" in workbook.sheetnames:
            sheet = workbook["rawdata_true_tx"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self._record_assignment(
                    assignments,
                    parse_asn(row[1] if len(row) > 1 else None),
                    source,
                    description=clean_text(row[2] if len(row) > 2 else None),
                    remark=clean_text(row[3] if len(row) > 3 else None),
                    domain="true-tx",
                )

        if "rawdata_true_CN" in workbook.sheetnames:
            sheet = workbook["rawdata_true_CN"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self._record_assignment(
                    assignments,
                    parse_asn(row[0] if len(row) > 0 else None),
                    source,
                    description=clean_text(row[1] if len(row) > 1 else None),
                    remark=clean_text(row[2] if len(row) > 2 else None),
                    domain="true-cn",
                )

    def _parse_external_workbook(
        self, workbook: Any, assignments: dict[int, dict[str, Any]], source: str
    ) -> None:
        sheet = workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            self._record_assignment(
                assignments,
                parse_asn(row[1] if len(row) > 1 else None),
                source,
                description=clean_text(row[2] if len(row) > 2 else None),
                remark=clean_text(row[3] if len(row) > 3 else None),
                site=clean_text(row[0] if len(row) > 0 else None),
                domain="external-true-it",
            )

    def _parse_aci_workbook(
        self, workbook: Any, assignments: dict[int, dict[str, Any]], source: str
    ) -> None:
        if "ASN" not in workbook.sheetnames:
            return

        sheet = workbook["ASN"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for asn_index, zone_index in ((0, 1), (4, 5), (8, 9)):
                zone = clean_text(row[zone_index] if len(row) > zone_index else None)
                if not zone:
                    continue
                self._record_assignment(
                    assignments,
                    parse_asn(row[asn_index] if len(row) > asn_index else None),
                    source,
                    description=zone,
                    domain=zone,
                )

    def _record_assignment(
        self,
        assignments: dict[int, dict[str, Any]],
        asn: int | None,
        source: str,
        *,
        description: str = "",
        remark: str = "",
        tenant: str = "",
        domain: str = "",
        site: str = "",
    ) -> None:
        if asn is None:
            return

        record = assignments.setdefault(
            asn,
            {
                "asn": asn,
                "sources": set(),
                "descriptions": set(),
                "remarks": set(),
                "tenants": set(),
                "domains": set(),
                "sites": set(),
            },
        )
        record["sources"].add(source)
        if description:
            record["descriptions"].add(description)
        if remark:
            record["remarks"].add(remark)
        if tenant:
            record["tenants"].add(tenant)
        if domain:
            record["domains"].add(domain)
        if site:
            record["sites"].add(site)

    def _serialize_assignment(self, record: dict[str, Any]) -> dict[str, Any]:
        return {
            "asn": record["asn"],
            "sources": sorted(record["sources"]),
            "descriptions": sorted(record["descriptions"]),
            "remarks": sorted(record["remarks"]),
            "tenants": sorted(record["tenants"]),
            "domains": sorted(record["domains"]),
            "sites": sorted(record["sites"]),
            "is_private": PRIVATE_ASN_START <= record["asn"] <= PRIVATE_ASN_END,
        }

    def _build_recommendations(
        self, blocks: list[dict[str, Any]], assignments: dict[int, dict[str, Any]]
    ) -> list[dict[str, Any]]:
        recommendations = []
        seen_asns: set[int] = set()
        for block in blocks:
            suggested = None
            for asn in range(block["start"], block["stop"] + 1):
                if asn not in assignments:
                    suggested = asn
                    break
            if suggested is None or suggested in seen_asns:
                continue
            seen_asns.add(suggested)
            recommendations.append(
                {
                    "block_name": block["name"],
                    "start": block["start"],
                    "stop": block["stop"],
                    "suggested_asn": suggested,
                }
            )
        return recommendations[:12]

    def _load_route_snapshot(self, path: Path) -> dict[str, Any]:
        lines: list[str] = []
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            for _, line in zip(range(120), handle):
                lines.append(line.rstrip())

        joined = "\n".join(lines)
        local_as = None
        router_id = None
        total_routes = None
        device_name = None
        vendor = "Unknown"

        nokia_match = re.search(r"BGP Router ID:(\S+)\s+AS:(\d+)", joined)
        if nokia_match:
            vendor = "Nokia"
            router_id = nokia_match.group(1)
            local_as = int(nokia_match.group(2))

        huawei_router_match = re.search(r"BGP Local router ID is (\S+)", joined)
        if huawei_router_match:
            vendor = "Huawei"
            router_id = huawei_router_match.group(1)

        huawei_total_match = re.search(r"Total number of routes from all PE:\s*(\d+)", joined)
        if huawei_total_match:
            total_routes = int(huawei_total_match.group(1))

        device_match = re.search(r"Device\s*:\s*(.+)", joined)
        if device_match:
            device_name = device_match.group(1).strip()
        elif lines:
            prompt_match = re.search(r"#\s*$", lines[0])
            if prompt_match:
                device_name = lines[0].split(":")[-1].replace("#", "").strip()

        return {
            "file_name": path.name,
            "vendor": vendor,
            "device_name": device_name or path.stem,
            "router_id": router_id,
            "local_as": local_as,
            "total_routes": total_routes,
            "size_mb": round(path.stat().st_size / (1024 * 1024), 2),
        }


def filter_assignments(assignments: list[dict[str, Any]], query: str, limit: int) -> list[dict[str, Any]]:
    normalized = query.strip().lower()
    if not normalized:
        return assignments[:limit]

    filtered = []
    for item in assignments:
        haystack = " ".join(
            [
                str(item["asn"]),
                " ".join(item["descriptions"]),
                " ".join(item["domains"]),
                " ".join(item["tenants"]),
                " ".join(item["sites"]),
                " ".join(item["sources"]),
            ]
        ).lower()
        if normalized in haystack:
            filtered.append(item)
        if len(filtered) >= limit:
            break
    return filtered
